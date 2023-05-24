import openpyxl

class WorkLocations:
    """Work locations to schedule employees to. Defines minimum and nominal number of employees per shift."""
    def __init__(self, location_name, location_address, shift_1_min_employees,
                 shift_1_nom_employees, shift_2_min_employees, shift_2_nom_employees,
                 shift_3_min_employees, shift_3_nom_employees):
        self.location_name = location_name
        self.location_address = location_address
        self.shift_1_min_employees = shift_1_min_employees
        self.shift_1_nom_employees = shift_1_nom_employees
        self.shift_2_min_employees = shift_2_min_employees
        self.shift_2_nom_employees = shift_2_nom_employees
        self.shift_3_min_employees = shift_3_min_employees
        self.shift_3_nom_employees = shift_3_nom_employees


    @classmethod
    def read_locations_from_xlsx(cls):
        """Reads WorkLocation data from xlsx file."""
        work_book = openpyxl.load_workbook('locations.xlsx')
        work_sheet = work_book.active
        locations = []
        for row in work_sheet.iter_rows(min_row=2, values_only=True):
            location = cls(*row)
            locations.append(location)
        return locations


    @classmethod
    def save_locations_to_xlsx(cls, locations):
        """Saves WorkLocation data to xlsx file after it was modified by user."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["location_name", "location_address", "shift_1_min_employees",
                   "shift_1_nom_employees", "shift_2_min_employees", "shift_2_nom_employees",
                   "shift_3_min_employees", "shift_3_nom_employees"])
        for location in locations:
            ws.append([location.location_name, location.location_address,
                       location.shift_1_min_employees, location.shift_1_nom_employees,
                       location.shift_2_min_employees, location.shift_2_nom_employees,
                       location.shift_3_min_employees, location.shift_3_nom_employees])
        wb.save('locations.xlsx')

    @classmethod
    def display_locations_info(cls):
        """Displays current WorkLocations instances."""
        locations = cls.read_locations_from_xlsx()
        for location in locations:
            print(f"Location Name: {location.location_name}")
            print(f"Location Address: {location.location_address}")
            print(f"Required employees: ")
            print(f"for shift 1 minimum {location.shift_1_min_employees}, nominal {location.shift_1_nom_employees}.")
            print(f"for shift 2 minimum {location.shift_2_min_employees}, nominal {location.shift_2_nom_employees}.")
            print(f"for shift 3 minimum {location.shift_3_min_employees}, nominal {location.shift_3_nom_employees}.")
           
            
if __name__ == '__main__':
    print(WorkLocations.read_locations_from_xlsx())
    WorkLocations.display_locations_info()
