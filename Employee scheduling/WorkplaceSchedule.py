import configparser
import datetime
from collections import defaultdict

import openpyxl
import pandas


class Settings:
    
    def __init__(self, settings_file:str) -> None:
        config = configparser.ConfigParser()
        config.read(settings_file)
        self.employee_file = config.get('employees', 'employee_data_file')
        self.location_file = config.get('locations', 'locations_data_file')
        self.schedule_range = (config.get('calendar', 'schedule_start_date'), config.get('calendar', 'schedule_stop_date'))
        self.weekends = ((True if config.get('calendar', 'working_saturdays').lower() in ('true', 'yes', '1', 'y', 't') else False, 
                          True if config.get('calendar', 'working_sundays').lower() in ('true', 'yes', '1', 'y', 't') else False))

class Employee:
    """Employee object. Contains employee_id, employee_name, employee_address, employee_schedule.
    Employee_schedule holds next shift time and workplace in dict format.
    """
    
    employee_list = []
    
    def __init__(self, employee_id:int = None, employee_name:str = None, employee_address:str = None, 
                 employee_schedule:dict = defaultdict()) -> None:
        
        self.employee_id = employee_id
        self.employee_name = employee_name
        self.employee_address = employee_address
        self.employee_schedule = defaultdict()
        if employee_schedule:
            self.employee_schedule.update(employee_schedule)
        
        
    @classmethod    
    def build_employee_list(cls, config:Settings) :
        """Reads employees from xlsx file and add them to employee_list."""  
        try:
            work_book = openpyxl.load_workbook(config.employee_file)
            work_sheet = work_book.active
            for row in work_sheet.iter_rows(min_row=2, values_only=True):                                        
                employee = cls(*row[:4])                                 # <-------- accepts first five arguments from xlsx file row.
                cls.employee_list.append(employee)
        except Exception as e:
            print(f"Exception occured while running {cls.build_employee_list.__name__}. Exception: {e}")
        return cls.employee_list
        
        
    @classmethod    
    def show_employees_schedules(cls) -> None:
        """Displays schedules of listed employees."""
        for employee in cls.employee_list:
            print(f'Employee id: {employee.employee_id}, {employee.employee_name}.')
            if not employee.employee_schedule:      
                print('Has nothing scheduled at the moment.')
                continue
            for scheduled in employee.employee_schedule:
                print(f'{scheduled} is working at {employee.employee_schedule[scheduled][0].workplace_name}. \
Workplace address is {employee.employee_schedule[scheduled][0].workplace_address}')          
        
        
class Workplace:
    """Workplace object. Contains workplace_id, workplace_name, workplace_address, min_employee, max_employee, workplace_schedule.
    Workplace_schedule holds next shift time and Employee objects list in dict format.
    """
    workplace_list = []
    
    def __init__(self, workplace_id:int = None, workplace_name:str = None, workplace_address:str = None, 
                 min_employee = 1, max_employee = 900, workplace_schedule:dict = defaultdict()) -> None:
        
        self.workplace_id = workplace_id
        self.workplace_name = workplace_name
        self.workplace_address = workplace_address
        if not isinstance(min_employee, int):               
            if not min_employee.isdigit():
                min_employee = 1
            else:
                min_employee = int(min_employee)
        self.min_employee = min_employee
        if not isinstance(max_employee, int):               
            if not max_employee.isdigit():
                max_employee = 1
            else:
                max_employee = int(max_employee)
        self.max_employee = max_employee
        self.workplace_schedule = defaultdict()
        if workplace_address:
            self.workplace_schedule.update(workplace_schedule)
        
        
    @classmethod    
    def build_workplace_list(cls, config:Settings) :
        """Reads workplaces from xlsx file and add them to workplace_list."""  
        try:
            work_book = openpyxl.load_workbook(config.location_file)
            work_sheet = work_book.active
            for row in work_sheet.iter_rows(min_row=2, values_only=True):                                        
                workplace = cls(*row[:4])                                     # <----- accepts first five arguments from xlsx file row. 
                cls.workplace_list.append(workplace)
        except Exception as e:
            print(f"Exception occured while running {cls.build_workplace_list.__name__}. Exception: {e}")
        return cls.workplace_list
               
            
    @classmethod    
    def show_workplace_schedules(cls) -> None:
        """Displays schedules for workplaces."""
        for workplace in cls.workplace_list:
            print(f'Workplace id: {workplace.workplace_id}, {workplace.workplace_name}.')
            if not workplace.workplace_schedule:      
                print('Has no employees scheduled at the moment.')
                continue
            for scheduled in workplace.workplace_schedule:
                print(f'At {scheduled} employees scheduled to work are:')
                for employee in workplace.workplace_schedule[scheduled]:
                    print(employee.employee_name, f'Employee Id: {employee.employee_id}')    
                       
            
class WorkdayCalendar:
    """WorkdayCalendar object holds list of shifts to be convered. Contains date, shift and schedule.
    Schedule holds shift information in Workplace object: Employee object list format."""
    
    shifts_to_schedule = []
    
    def __init__(self, date:datetime, shift:int = 1, schedule:dict = defaultdict()) -> None:
        self.date = date
        
        self.shift = shift
        self.schedule = defaultdict()
        if schedule:
            self.schedule.update(schedule)
        
    @classmethod
    def initialize_schedule(cls, config, employee_list, workplace_list) -> None:
        """Build schedule assignments. 
        WorkdayCalendar.schedule will append dict with Workplace: Employee.
        Workplace.workplace_schedule will append dict with date + hour: Employee.
        Employee.employee_schedule will append dict with date + hour: Workplace.
        Will create unassigned employees objects as WorkdayCalendar.schedule['unassigned'].        
        """
        #Get schedule requirements from config file.
        start_date = config.schedule_range[0]
        end_date = config.schedule_range[1]
        min_employees = sum(list(map(lambda x: x.min_employee*3 , workplace_list)))     # <--------- assumes three shifts.
        max_employees = sum(list(map(lambda x: x.max_employee*3 , workplace_list)))     # <--------- assumes three shifts.
        shift_offset = {1: datetime.timedelta(hours = 6), 2: datetime.timedelta(hours = 14), 3: datetime.timedelta(hours = 22) }
        
        #Check if there is enough employees to build a schedule. 
        if min_employees > len(employee_list):          
            print(f'''There are not enough employees for minimum requirements of all workplaces for three shifts. 
                  You need additional {len(employee_list)-min_employees} employees''')      
        if max_employees < len(employee_list): 
            print(f'There are more employees than available workspaces. There will be {max_employees-len(employee_list)} employees unassigned.')
        else:
            print(f'There are {len(employee_list)} employees and {max_employees} workspaces. \
                Minimum required employees: {min_employees} Employees will be assigned to available spaces.' )
       
       #Create list of days to cover.
        for day in pandas.date_range(start_date, end_date):                             # check if its sat or sun and compare with settings. 
            if day.weekday() == 5 and not config.weekends[0]:
                continue
            if day.weekday() == 6 and not config.weekends[1]:  
                continue
            for shift in range(1,4):
                cls.shifts_to_schedule.append(WorkdayCalendar(day, shift))

        #Build assignments for each shift.
        previous_day = datetime.datetime.min
        remaining_employees = defaultdict(list)
        for day_shift in cls.shifts_to_schedule:
            #refresh available employees list. 
            if not previous_day == day_shift.date:
                available_employees = employee_list.copy()
                available_employees.sort(key= cls._for_sort_employee_list, reverse = True)    
                previous_day = day_shift.date      
            workplace_fill_min = workplace_list.copy()
            workplace_len, employee_len = 0, 0
            while workplace_fill_min:
                # Infinite loop prevention in case of unassignable employees. 
                if (workplace_len, employee_len) == (len(workplace_fill_min), len(available_employees)):
                    break
                else:
                    (workplace_len, employee_len) = (len(workplace_fill_min), len(available_employees))    
                if not available_employees:
                    break
                for workplace in workplace_fill_min:
                    if workplace.min_employee <= len(workplace.workplace_schedule.setdefault(day_shift.date + shift_offset[day_shift.shift], []) ):
                        workplace_fill_min.remove(workplace)
                        continue
                    #Check if employee had at least 11 hours break. 
                    for employee in available_employees:
                        if employee.employee_schedule and list(employee.employee_schedule.keys())[-1] + datetime.timedelta(hours=19) \
                                                            >= day_shift.date + shift_offset[day_shift.shift]:
                            continue
                        #assign employee and remove from list.
                        cls._assign_schedule(day_shift, employee, workplace)
                        available_employees.remove(employee)
                        break
            if available_employees and day_shift.shift == 3:
                remaining_employees[day_shift.date].extend(available_employees)
        
        for day_shift in cls.shifts_to_schedule:
            workplace_fill_max = workplace_list.copy()
            
            workplace_len, employee_len = 0, 0
            while workplace_fill_max and remaining_employees[day_shift.date]:
                # Infinite loop prevention in case of unassignable employees. 
                if (workplace_len, employee_len) == (len(workplace_fill_max), len(remaining_employees[day_shift.date])):
                    break
                else:
                    (workplace_len, employee_len) = (len(workplace_fill_max), len(remaining_employees[day_shift.date]))    
                remaining_employees[day_shift.date].sort(key= cls._for_sort_employee_list, reverse = True)
                # Find employee for workplace.
                for workplace in workplace_fill_max:  
                    if workplace.max_employee <= len(workplace.workplace_schedule.setdefault(day_shift.date + shift_offset[day_shift.shift], []) ):
                        workplace_fill_max.remove(workplace)
                        continue         
                    for employee in remaining_employees[day_shift.date]:
                        if employee.employee_schedule and list(employee.employee_schedule.keys())[-1] + datetime.timedelta(hours=19) \
                                                            >= day_shift.date + shift_offset[day_shift.shift]:
                            continue
                        #assign employee and remove from list.
                        cls._assign_schedule(day_shift, employee, workplace)
                        remaining_employees[day_shift.date].remove(employee)
                        break
 
                                             
    @staticmethod                        
    def _for_sort_employee_list(employee) -> datetime:
        """Internal function to sort employees based on last shift."""
        if not employee.employee_schedule:
            return datetime.datetime.min
        return list(employee.employee_schedule.keys())[-1]
    
    
    def _assign_schedule(self, employee:object, workplace:object) -> bool:
        """Assigns schedule. 
        Date shift dict updates dict with workplace: employee.
        Workplace shift dict updates with date, shift: employee
        Employee shift dict updates with date: workplace.
        Returns True on successful schedule assignment. 
        """
        shift_offset = {1: datetime.timedelta(hours = 6), 2: datetime.timedelta(hours = 14), 3: datetime.timedelta(hours = 22) }
        #Assign date: workplace to employee.
        if not employee.employee_schedule or not employee.employee_schedule.setdefault(self.date + shift_offset[self.shift], []):
            employee.employee_schedule[self.date + shift_offset[self.shift]] = [workplace]
            # print(f'Employee! Assigned {employee.employee_name}, {employee.employee_id} to {workplace.workplace_name} at {self.date}, shift: {self.shift}.')
            # print(f'Proof: {employee.employee_schedule[self.date + shift_offset[self.shift]]} ')         # Print debugging
        else:
            print(f'something went wrong with schedule of employee {employee.employee_name}, {employee.employee_id}. \
                  Tried to schedule employee to {self.date}, shift: {self.shift} at {workplace.workplace_name} but {employee.employee_name} \
                  already has schedule for {self.date + shift_offset[self.shift]}: {employee.employee_schedule[self.date + shift_offset[self.shift]]}.') 
            return False       
        #Assign workplace: employee to date with shift.
        if not self.schedule or not self.schedule.setdefault(workplace, []):
            self.schedule[workplace] = [employee]
            # print('!!',self.date, self.shift, self.schedule[workplace][-1])           # Print debugging
        else:
            self.schedule[workplace].append(employee)
            # print('??',self.date, self.shift, self.schedule[workplace][-1])           # Print debugging
        # print(f'Shift! Assigned {employee.employee_name}, {employee.employee_id} to {workplace.workplace_name} at {self.date}, shift: {self.shift}.')
        # print(f'Proof: {self.schedule[workplace]} ')                                                        # Print debugging
        #Assign date: employee to workplace.        
        if not workplace.workplace_schedule or not workplace.workplace_schedule[self.date + shift_offset[self.shift]]:   
            workplace.workplace_schedule[self.date + shift_offset[self.shift]] = [employee]   
        else:
            workplace.workplace_schedule[self.date + shift_offset[self.shift]].append(employee)
        # print(f'Workplace! Assigned {employee.employee_name}, {employee.employee_id} to {workplace.workplace_name} at {self.date}, shift: {self.shift}.')
        # print(f'Proof: {workplace.workplace_schedule[self.date + shift_offset[self.shift]]} ')              # Print debugging
        return True


    @classmethod
    def show_all_schedules(cls) -> None:
        """Shows all schedules by date and shift"""
        if not cls.shifts_to_schedule:
            print('No shifts planned yet.')
        for shift in cls.shifts_to_schedule:
            if not shift.schedule:
                print(f'No shifts planned yet for {shift.date}, shift {shift.shift}.')
                
            print(f'Shifts planned for {shift.date} shift {shift.shift}:')
            for workplace in shift.schedule:
                print(f'Workplace {workplace.workplace_name}, id: {workplace.workplace_id} has scheduled employees:')
                for employee in shift.schedule[workplace]:
                    print(f'{employee.employee_name} id: {employee.employee_id}')
        

def main(settings):
    settings = Settings(settings)
    Employee.build_employee_list(settings)
    Workplace.build_workplace_list(settings)
    WorkdayCalendar.initialize_schedule(settings, Employee.employee_list, Workplace.workplace_list)
    return settings
    
if __name__ == '__main__':
    settings = main('settings.ini')
    
    # Employee.show_employees_schedules()
    # Workplace.show_workplace_schedules()
    
    
    WorkdayCalendar.show_all_schedules()